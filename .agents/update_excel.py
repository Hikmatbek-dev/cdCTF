import openpyxl
import os

file_path = "Стартаплар рўйҳати.xlsx"

if not os.path.exists(file_path):
    print("File not found")
    exit(1)

wb = openpyxl.load_workbook(file_path)
ws = wb.active

maqsad = "O'zbekiston yoshlari uchun 3 tilli (o'zbek, rus, ingliz) innovatsion kiberxavfsizlik akademiyasi va amaliy CTF platformasini yaratish. Yoshlarni noldan o'qitib, nufuzli ishlarga joylashishlari uchun ko'prik vazifasini o'tash."
holat = "MVP to'liq ishga tushirilgan (cdctf.uz). Hozirda platformada 8 ta ta'lim moduli, 64 ta dars va 97 ta amaliy CTF topshirig'i mavjud bo'lib, ilk foydalanuvchilar muvaffaqiyatli jalb etilgan."
sotuvlar = "Loyiha ayni paytda faol foydalanuvchilar bazasini shakllantirish (Traction) bosqichida. Kelgusida B2B (kadrlar yetkazib berish, korporativ ta'lim) va homiylik musobaqalari orqali to'liq monetizatsiya qilinadi."
muammolar = "Platformani keng ommaga olib chiqish uchun marketing byudjetining yo'qligi; amaliy laboratoriyalar uchun bulutli serverlar infratuzilmasi xarajatlari; malakali kadrlarni jamoaga jalb qilish uchun sarmoya yetishmovchiligi."
invest_miqdor = "10,000 AQSh dollari (10 oylik operatsion xarajatlar uchun)"
invest_maqsad = "10 oy davomida jadal o'sishni ta'minlash maqsadida: amaliy laboratoriyalar serverlarini saqlab turish, jamoani qo'llab-quvvatlash hamda foydalanuvchilarni jalb etish uchun maqsadli marketing kampaniyalarini o'tkazishga sarflanadi."
oldin_invest = "Yo'q. Loyiha shu kungacha to'liq asoschilarning o'z mablag'lari va shaxsiy vaqti (Bootstrapping) hisobiga qurilgan."

target_row = 4

ws.cell(row=target_row, column=1, value=1)
ws.cell(row=target_row, column=2, value="Hikmatbek Xudoyberganov")
ws.cell(row=target_row, column=3, value="CdCTF")
ws.cell(row=target_row, column=4, value=maqsad)
ws.cell(row=target_row, column=5, value=holat)
ws.cell(row=target_row, column=6, value=3)
ws.cell(row=target_row, column=7, value=sotuvlar)
ws.cell(row=target_row, column=8, value=muammolar)
ws.cell(row=target_row, column=9, value=invest_miqdor)
ws.cell(row=target_row, column=10, value=invest_maqsad)
ws.cell(row=target_row, column=11, value=oldin_invest)

wb.save(file_path)
print(f"Successfully refined and updated row {target_row} for investors")
